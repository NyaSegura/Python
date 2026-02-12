import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

def load_xyz_txt(path: str) -> pd.DataFrame:
    # Reads space-separated x y z points from a TXT file
    df = pd.read_csv(path, sep=r"\s+", header=None, names=["x", "y", "z"])
    if df.shape[1] < 3:
        raise ValueError("TXT file must contain at least three columns for x, y, z coordinates.")
    df = df.iloc[:, :3].copy()
    df.columns = ["x", "y", "z"]
    return df

def write_column(ws, col_letter: str, start_row: int, values):
    for i, value in enumerate(values, start=start_row):
        ws[f"{col_letter}{i}"] = float(value)


def get_app_directory():
    if getattr(sys, 'frozen', False):
        # Running as EXE
        return os.path.dirname(sys.executable)
    else:
        # Running as script
        return os.path.dirname(os.path.abspath(__file__))
    
APP_DIR = get_app_directory()

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("TTV Template Filler")
        self.geometry("720x420")
        self.resizable(False, False)

        # Defaults (edit once you know your true template settings)

        self.template_path = tk.StringVar(value=os.path.join(APP_DIR, "TTV_template.xlsx"))
        self.side1_path = tk.StringVar(value=os.path.join(APP_DIR, "TTV_side1.txt"))
        self.side2_path = tk.StringVar(value=os.path.join(APP_DIR, "TTV_side2.txt"))
        self.output_path = tk.StringVar(value=os.path.join(APP_DIR, "TTV_output.xlsx"))

        self.sheet1_name = tk.StringVar(value="Side 1")
        self.sheet2_name = tk.StringVar(value="Side 2")
        self.column_letter = tk.StringVar(value="D")
        self.start_row = tk.StringVar(value="3")


        pad = {"padx": 10, "pady": 6}

        def browse_file(var: tk.StringVar, types):
            path = filedialog.askopenfilename(filetypes=types)
            if path:
                var.set(path)

        def browse_save(var: tk.StringVar):
            path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save output as...")
            if path:
                var.set(path)

        # Inputs
        frm = tk.Frame(self)
        frm.pack(fill="both", expand=True, padx=10, pady=10)

        # Template file
        tk.Label(frm, text="Excel Template File:").grid(row=0, column=0, sticky="w", **pad)
        tk.Entry(frm, textvariable=self.template_path, width=70).grid(row=0, column=1, **pad)
        tk.Button(frm, text="Browse...", command=lambda: browse_file(self.template_path, [("Excel Files", "*.xlsx")])).grid(row=0, column=2, **pad)

        # Side 1 TXT file
        tk.Label(frm, text="Side 1 TXT File:").grid(row=1, column=0, sticky="w", **pad)
        tk.Entry(frm, textvariable=self.side1_path, width=70).grid(row=1, column=1, **pad)
        tk.Button(frm, text="Browse...", command=lambda: browse_file(self.side1_path, [("Text Files", "*.txt"), ("All Files", "*.*")])).grid(row=1, column=2, **pad)

        # Side 2 TXT file
        tk.Label(frm, text="Side 2 TXT File:").grid(row=2, column=0, sticky="w", **pad)
        tk.Entry(frm, textvariable=self.side2_path, width=70).grid(row=2, column=1, **pad)
        tk.Button(frm, text="Browse...", command=lambda: browse_file(self.side2_path, [("Text Files", "*.txt"), ("All Files", "*.*")])).grid(row=2, column=2, **pad)

        # Output file
        tk.Label(frm, text="Output Excel File:").grid(row=3, column=0, sticky="w", **pad)
        tk.Entry(frm, textvariable=self.output_path, width=70).grid(row=3, column=1, **pad)
        tk.Button(frm, text="Save As", command=lambda: browse_save(self.output_path)).grid(row=3, column=2, **pad)
        
        # Settings box
        box = tk.LabelFrame(frm, text="Template Settings")
        box.grid(row=4, column=0, columnspan=3, sticky="we", padx=10, pady=14)

        tk.Label(box, text="Sheet name for Side 1:").grid(row=0, column=0, sticky="w", padx=10, pady=6)
        tk.Entry(box, textvariable=self.sheet1_name, width=18).grid(row=0, column=1, sticky="w", padx=10, pady=6)

        tk.Label(box, text="Sheet name for Side 2:").grid(row=0, column=2, sticky="w", padx=10, pady=6)
        tk.Entry(box, textvariable=self.sheet2_name, width=18).grid(row=0, column=3, sticky="w", padx=10, pady=6)

        tk.Label(box, text="Write Z into column:").grid(row=1, column=0, sticky="w", padx=10, pady=6)
        tk.Entry(box, textvariable=self.column_letter, width=6).grid(row=1, column=1, sticky="w", padx=10, pady=6)

        tk.Label(box, text="Start row:").grid(row=1, column=2, sticky="w", padx=10, pady=6)
        tk.Entry(box, textvariable=self.start_row, width=6).grid(row=1, column=3, sticky="w", padx=10, pady=6)

        # Run button
        tk.Button(frm, text="RUN", height=2, command=self.run).grid(row=5, column=0, columnspan=3, pady=10)

        # Status
        self.status = tk.StringVar(value="Ready.")
        tk.Label(frm, textvariable=self.status, anchor="w").grid(row=6, column=0, columnspan=3, sticky="we", padx=10, pady=6)

    def run(self):
        
        messagebox.showinfo("Debug", "RUN clicked — starting process.")
        
        try:
            template = self.template_path.get().strip()
            s1 = self.side1_path.get().strip()
            s2 = self.side2_path.get().strip()
            out = self.output_path.get().strip()

            if not (template and s1 and s2 and out):
                messagebox.showerror("Missing input", "Please select template, Side 1, Side 2, and an output file.")
                return

            sheet1 = self.sheet1_name.get().strip()
            sheet2 = self.sheet2_name.get().strip()
            col = str(self.column_letter.get()).strip().upper()

            try:
                start_row = int(str(self.start_row.get()).strip())
            except ValueError:
                messagebox.showerror("Invalid Start Row", "Start row must be an integer.")
                return



            self.status.set("Loading TXT files...")
            self.update_idletasks()

            side1 = load_xyz_txt(s1)
            side2 = load_xyz_txt(s2)

            if len(side1) != len(side2):
                # Not necessarily fatal for your workflow, but usually suspicious
                if not messagebox.askyesno(
                    "Point count mismatch",
                    f"Side 1 has {len(side1)} points, Side 2 has {len(side2)} points.\n"
                    "Continue anyway?"
                ):
                    self.status.set("Cancelled.")
                    return

            self.status.set("Opening Excel template...")
            self.update_idletasks()

            wb = load_workbook(template)

            if sheet1 not in wb.sheetnames or sheet2 not in wb.sheetnames:
                raise ValueError(
                    f"Sheet not found. Available sheets: {', '.join(wb.sheetnames)}"
                )

            ws1 = wb[sheet1]
            ws2 = wb[sheet2]

            self.status.set("Writing Z columns...")
            self.update_idletasks()

            write_column(ws1, col, start_row, side1["z"])
            write_column(ws2, col, start_row, side2["z"])

            self.status.set("Saving output...")
            self.update_idletasks()

            wb.save(out)

            self.status.set("Done.")
            messagebox.showinfo("Success", f"Saved filled template:\n{out}")

        except Exception as e:
            self.status.set("Error.")
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    App().mainloop()
        